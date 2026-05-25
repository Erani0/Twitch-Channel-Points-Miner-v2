import json
import os
import tempfile
import time
import unittest
from datetime import datetime, timezone
from unittest.mock import Mock, patch

from openpyxl import load_workbook

from TwitchChannelPointsMiner.TwitchChannelPointsMiner import TwitchChannelPointsMiner
from TwitchChannelPointsMiner.WatchStreakCache import WatchStreakCache
from TwitchChannelPointsMiner.classes.entities.Streamer import Streamer
from TwitchChannelPointsMiner.utils import _millify


class StreamersExportTest(unittest.TestCase):
    def _make_miner(self, export_path: str) -> TwitchChannelPointsMiner:
        miner = TwitchChannelPointsMiner.__new__(TwitchChannelPointsMiner)
        miner.streamers = []
        miner.original_streamers = []
        miner.daily_points_baseline_path = os.path.join(
            os.path.dirname(export_path) or ".",
            "daily_points_baseline.tester.json",
        )
        miner.daily_points_day_key = datetime.now().strftime("%Y-%m-%d")
        miner.daily_points_baseline = {}
        miner.daily_points_snapshot = {}
        miner.daily_points_session_anchor = {}
        miner._daily_points_baseline_dirty = False
        miner.streamer_follow_dates = {}
        miner.streamers_export_path = export_path
        miner.streamers_export_thread = None
        miner.streamers_export_interval_seconds = 600
        miner.running = False
        miner.username = "tester"
        miner.watch_streak_cache = WatchStreakCache(default_account_name="tester")
        miner._watch_streak_days_lookup_attempted = set()
        miner._chat_ban_lookup_attempted = set()
        return miner

    def test_build_streamer_export_rows_sorted_and_formatted(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            miner = self._make_miner(os.path.join(tmp_dir, "streamers.xlsx"))

            easyemi = Streamer("easyemi")
            easyemi.channel_points = 184880
            easyemi.subscription_tier = 1
            easyemi.chat_banned = True
            easyemi.stream.created_at = datetime(
                2026, 3, 4, tzinfo=timezone.utc
            ).timestamp()
            easyemi.history = {
                "WATCH": {"counter": 2, "amount": 20},
                "WATCH_STREAK": {"counter": 1, "amount": 450},
            }

            rubia = Streamer("rubia")
            rubia.channel_points = 26870
            rubia.subscription_tier = 1
            rubia.stream.created_at = datetime(
                2026, 3, 5, tzinfo=timezone.utc
            ).timestamp()
            rubia.history = {"CLAIM": {"counter": 1, "amount": 999}}

            itsceydi = Streamer("itsceydi")
            itsceydi.channel_points = 82570
            itsceydi.subscription_tier = None
            itsceydi.chat_banned = False
            itsceydi.stream.created_at = datetime(
                2026, 3, 6, tzinfo=timezone.utc
            ).timestamp()
            itsceydi.history = {}

            miner.streamers = [rubia, easyemi, itsceydi]
            miner.streamer_follow_dates = {
                "easyemi": "2025-07-21T12:34:56Z",
                "rubia": None,
            }
            miner.watch_streak_cache = WatchStreakCache(default_account_name="tester")
            miner.watch_streak_cache.ensure_session("easyemi", "broadcast-a", 1)
            miner.watch_streak_cache.mark_claimed("easyemi", "broadcast-a", 2)
            miner.watch_streak_cache.ensure_session("easyemi", "broadcast-b", 3)
            miner.watch_streak_cache.mark_claimed("easyemi", "broadcast-b", 4)
            miner.watch_streak_cache.ensure_session("rubia", "broadcast-c", 5)
            miner.watch_streak_cache.set_streamer_status(
                "easyemi",
                watch_streak_detected=True,
                watch_streak_days=12,
                is_online=True,
                last_stream_started_at=easyemi.stream.created_at,
                broadcast_id="broadcast-b",
                checked_at=easyemi.stream.created_at,
                account_name="tester",
            )
            miner.daily_points_baseline = {
                "easyemi": 0,
                "itsceydi": 0,
                "rubia": 0,
            }
            miner.daily_points_session_anchor = {
                "easyemi": 0,
                "itsceydi": 0,
                "rubia": 0,
            }

            rows = miner._build_streamer_export_rows()

            self.assertEqual([row["Streamer"] for row in rows], ["easyemi", "itsceydi", "rubia"])
            self.assertEqual(rows[0]["Points"], _millify(184880))
            self.assertEqual(rows[1]["Points"], _millify(82570))
            self.assertEqual(rows[2]["Points"], _millify(26870))
            self.assertEqual(rows[0]["Followdate"], "21.07.2025")
            self.assertEqual(rows[1]["Followdate"], "...")
            self.assertEqual(rows[2]["Followdate"], "...")
            self.assertEqual(rows[0]["Last Stream"], "04.03.2026")
            self.assertEqual(rows[1]["Last Stream"], "06.03.2026")
            self.assertEqual(rows[2]["Last Stream"], "05.03.2026")
            self.assertEqual(rows[0]["Sub"], "yes")
            self.assertEqual(rows[1]["Sub"], "no")
            self.assertEqual(rows[2]["Sub"], "yes")
            self.assertEqual(rows[0]["Banned"], "yes")
            self.assertEqual(rows[1]["Banned"], "no")
            self.assertEqual(rows[2]["Banned"], "no")
            self.assertEqual(rows[0]["Watchstreaks"], 12)
            self.assertEqual(rows[1]["Watchstreaks"], 0)
            self.assertEqual(rows[2]["Watchstreaks"], 0)
            self.assertEqual(rows[0]["Points gained"], 470)
            self.assertEqual(rows[1]["Points gained"], 0)
            self.assertEqual(rows[2]["Points gained"], 999)

    def test_points_gained_daily_starts_from_zero_without_baseline(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            miner = self._make_miner(os.path.join(tmp_dir, "streamers.xlsx"))

            streamer = Streamer("demo")
            streamer.channel_points = 1000
            streamer.history = {
                "WATCH": {"counter": 1, "amount": 12},
                "CLAIM": {"counter": 1, "amount": 50},
            }

            miner.streamers = [streamer]

            rows = miner._build_streamer_export_rows()
            self.assertEqual(rows[0]["Points gained"], 0)

    def test_points_gained_persists_across_same_day_restart(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            export_path = os.path.join(tmp_dir, "streamers.xlsx")
            miner = self._make_miner(export_path)

            first_streamer = Streamer("demo")
            first_streamer.history = {"WATCH": {"counter": 1, "amount": 120}}
            miner.streamers = [first_streamer]

            first_rows = miner._build_streamer_export_rows()
            self.assertEqual(first_rows[0]["Points gained"], 0)
            miner._save_daily_points_baseline_if_dirty()

            first_streamer.history = {
                "WATCH": {"counter": 1, "amount": 120},
                "CLAIM": {"counter": 1, "amount": 50},
            }
            updated_rows = miner._build_streamer_export_rows()
            self.assertEqual(updated_rows[0]["Points gained"], 50)
            miner._save_daily_points_baseline_if_dirty()

            restarted = self._make_miner(export_path)
            restarted._load_daily_points_baseline()

            second_streamer = Streamer("demo")
            second_streamer.history = {}
            restarted.streamers = [second_streamer]

            second_rows = restarted._build_streamer_export_rows()
            self.assertEqual(second_rows[0]["Points gained"], 50)

    def test_points_gained_ignores_legacy_baseline_payload(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            export_path = os.path.join(tmp_dir, "streamers.xlsx")
            miner = self._make_miner(export_path)

            with open(miner.daily_points_baseline_path, "w", encoding="utf-8") as handle:
                json.dump(
                    {
                        "day_key": datetime.now().strftime("%Y-%m-%d"),
                        "baseline": {"demo": 999},
                    },
                    handle,
                )

            miner._load_daily_points_baseline()
            self.assertEqual(miner.daily_points_baseline, {})
            self.assertEqual(miner.daily_points_snapshot, {})

            streamer = Streamer("demo")
            streamer.history = {}
            miner.streamers = [streamer]

            rows = miner._build_streamer_export_rows()
            self.assertEqual(rows[0]["Points gained"], 0)

    def test_points_gained_daily_resets_on_day_change(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            miner = self._make_miner(os.path.join(tmp_dir, "streamers.xlsx"))

            streamer = Streamer("demo")
            streamer.history = {"WATCH": {"counter": 1, "amount": 120}}

            miner.streamers = [streamer]
            miner.daily_points_day_key = "2000-01-01"
            miner.daily_points_baseline = {"demo": 0}

            rows = miner._build_streamer_export_rows()
            self.assertEqual(rows[0]["Points gained"], 0)

    def test_watch_streak_days_fetches_twitch_value_when_cache_is_missing_days(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            miner = self._make_miner(os.path.join(tmp_dir, "streamers.xlsx"))

            streamer = Streamer("demo")
            streamer.channel_id = "123456"
            miner.streamers = [streamer]
            miner.watch_streak_cache.ensure_session(
                streamer.username,
                "broadcast-a",
                1,
                account_name="tester",
            )
            miner.watch_streak_cache.mark_claimed(
                streamer.username,
                "broadcast-a",
                2,
                account_name="tester",
            )
            miner.watch_streak_cache.ensure_session(
                streamer.username,
                "broadcast-b",
                3,
                account_name="tester",
            )
            miner.watch_streak_cache.mark_claimed(
                streamer.username,
                "broadcast-b",
                4,
                account_name="tester",
            )
            miner.twitch = Mock()
            miner.twitch.get_watch_streak_days.return_value = 26

            rows = miner._build_streamer_export_rows()

            self.assertEqual(rows[0]["Watchstreaks"], 26)
            miner.twitch.get_watch_streak_days.assert_called_once_with(streamer)
            status = miner.watch_streak_cache.get_streamer_status(
                streamer.username,
                account_name="tester",
            )
            self.assertIsNotNone(status)
            self.assertEqual(status.watch_streak_days, 26)

    def test_banned_column_fetches_twitch_value_when_startup_state_is_unknown(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            miner = self._make_miner(os.path.join(tmp_dir, "streamers.xlsx"))

            streamer = Streamer("demo")
            streamer.channel_id = "123456"
            miner.streamers = [streamer]
            miner.twitch = Mock()
            miner.twitch.get_chat_ban_status.return_value = True

            rows = miner._build_streamer_export_rows()

            self.assertEqual(rows[0]["Banned"], "yes")
            miner.twitch.get_chat_ban_status.assert_called_once_with(streamer)
            self.assertTrue(streamer.chat_banned)

    @unittest.skipUnless(hasattr(time, "tzset"), "timezone switching requires tzset")
    def test_format_timestamp_date_uses_utc_date(self):
        miner = self._make_miner("streamers.xlsx")
        timestamp = datetime(2026, 3, 4, 0, 30, tzinfo=timezone.utc).timestamp()
        original_tz = os.environ.get("TZ")

        try:
            os.environ["TZ"] = "America/New_York"
            time.tzset()
            self.assertEqual(miner._format_timestamp_date(timestamp), "04.03.2026")
        finally:
            if original_tz is None:
                os.environ.pop("TZ", None)
            else:
                os.environ["TZ"] = original_tz
            time.tzset()

    def test_last_stream_date_prefers_latest_start_timestamp(self):
        miner = self._make_miner("streamers.xlsx")
        streamer = Streamer("demo")
        started_at = datetime(2026, 3, 5, 10, 0, tzinfo=timezone.utc).timestamp()
        ended_at = datetime(2026, 3, 6, 12, 0, tzinfo=timezone.utc).timestamp()

        miner.watch_streak_cache.ensure_session(
            streamer.username,
            "broadcast-demo",
            started_at,
            account_name="tester",
        )
        miner.watch_streak_cache.mark_ended(
            streamer.username,
            "broadcast-demo",
            ended_at=ended_at,
            account_name="tester",
        )

        self.assertEqual(miner._last_stream_date(streamer), "05.03.2026")

    def test_write_streamers_xlsx_applies_workbook_formatting(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            export_path = os.path.join(tmp_dir, "streamers.xlsx")
            miner = self._make_miner(export_path)

            rows = [
                {
                    "Streamer": "very_long_streamer_name",
                    "Points": "184.88k",
                    "Followdate": "21.07.2025",
                    "Last Stream": "04.03.2026",
                    "Sub": "yes",
                    "Banned": "yes",
                    "Watchstreaks": 26,
                    "Points gained": 460,
                }
            ]

            miner._write_streamers_xlsx(rows)

            self.assertTrue(os.path.isfile(export_path))
            workbook = load_workbook(export_path)
            sheet = workbook.active

            for header in ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]:
                self.assertTrue(sheet[header].font.bold)
                self.assertEqual(sheet[header].alignment.horizontal, "center")
                self.assertEqual(sheet[header].fill.fill_type, "solid")
                self.assertTrue((sheet[header].fill.fgColor.rgb or "").endswith("1F4E78"))
                self.assertIsNone(sheet[header].border.bottom.style)

            self.assertEqual(sheet.title, "Sheet1")
            self.assertIsNone(sheet.sheet_view.showGridLines)
            self.assertIsNone(sheet.sheet_view.zoomScale)
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual(sheet.row_dimensions[1].height, 22)
            self.assertIsNone(sheet.row_dimensions[2].height)
            self.assertEqual(len(sheet.tables), 1)

            table = next(iter(sheet.tables.values()))
            self.assertEqual(table.ref, "A1:H2")
            self.assertEqual(table.tableStyleInfo.name, "TableStyleMedium2")
            self.assertTrue(table.tableStyleInfo.showRowStripes)
            self.assertEqual(table.autoFilter.ref, "A1:H2")

            self.assertIsNone(sheet["A2"].hyperlink)
            self.assertIsNone(sheet["A2"].alignment.horizontal)
            self.assertEqual(sheet["B2"].alignment.horizontal, "right")
            self.assertEqual(sheet["C2"].alignment.horizontal, "center")
            self.assertEqual(sheet["D2"].alignment.horizontal, "center")
            self.assertEqual(sheet["E2"].alignment.horizontal, "center")
            self.assertEqual(sheet["F2"].alignment.horizontal, "center")
            self.assertEqual(sheet["G2"].alignment.horizontal, "center")
            self.assertEqual(sheet["H2"].alignment.horizontal, "right")

            self.assertEqual(sheet["C2"].number_format, "DD.MM.YYYY")
            self.assertEqual(sheet["D2"].number_format, "DD.MM.YYYY")
            self.assertEqual(sheet["C2"].value.strftime("%d.%m.%Y"), "21.07.2025")
            self.assertEqual(sheet["D2"].value.strftime("%d.%m.%Y"), "04.03.2026")

            self.assertIsNone(sheet["E2"].fill.fill_type)
            self.assertIsNone(sheet["F2"].fill.fill_type)
            self.assertTrue((sheet["G2"].fill.fgColor.rgb or "").endswith("E8F5E9"))
            self.assertTrue((sheet["H2"].fill.fgColor.rgb or "").endswith("F1F8E9"))
            self.assertTrue(
                any(str(entry.sqref).startswith("H2") for entry in sheet.conditional_formatting)
            )

            self.assertEqual(sheet.column_dimensions["A"].width, 20)
            self.assertEqual(sheet.column_dimensions["B"].width, 10)
            self.assertEqual(sheet.column_dimensions["C"].width, 13)
            self.assertEqual(sheet.column_dimensions["D"].width, 13)
            self.assertEqual(sheet.column_dimensions["E"].width, 8)
            self.assertEqual(sheet.column_dimensions["F"].width, 10)
            self.assertEqual(sheet.column_dimensions["G"].width, 14)
            self.assertEqual(sheet.column_dimensions["H"].width, 15)

    def test_write_streamers_xlsx_rolls_to_new_date_file(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            old_export_path = os.path.join(tmp_dir, "report_2026-03-05_tester.xlsx")
            new_export_path = os.path.join(tmp_dir, "report_2026-03-06_tester.xlsx")
            miner = self._make_miner(old_export_path)

            rows = [
                {
                    "Streamer": "easyemi",
                    "Points": "184.88k",
                    "Followdate": "21.07.2025",
                    "Last Stream": "04.03.2026",
                    "Sub": "yes",
                    "Banned": "yes",
                    "Watchstreaks": 26,
                    "Points gained": 460,
                }
            ]

            with patch.object(
                TwitchChannelPointsMiner,
                "_current_streamers_export_path",
                return_value=new_export_path,
            ):
                miner._write_streamers_xlsx(rows)

            self.assertEqual(miner.streamers_export_path, new_export_path)
            self.assertFalse(os.path.isfile(old_export_path))
            self.assertTrue(os.path.isfile(new_export_path))

    def test_write_streamers_xlsx_skips_points_gained_databar_when_all_zero(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            export_path = os.path.join(tmp_dir, "streamers.xlsx")
            miner = self._make_miner(export_path)

            rows = [
                {
                    "Streamer": "demo",
                    "Points": "1.20k",
                    "Followdate": "21.07.2025",
                    "Last Stream": "04.03.2026",
                    "Sub": "no",
                    "Banned": "no",
                    "Watchstreaks": 0,
                    "Points gained": 0,
                }
            ]

            miner._write_streamers_xlsx(rows)

            workbook = load_workbook(export_path)
            sheet = workbook.active
            self.assertEqual(len(sheet.conditional_formatting), 0)

    def test_streamers_export_loop_runs_periodic_export(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            miner = self._make_miner(os.path.join(tmp_dir, "streamers.xlsx"))
            miner.running = True

            calls = []

            def fake_export():
                calls.append(1)
                miner.running = False

            with patch.object(
                TwitchChannelPointsMiner,
                "_export_streamers_snapshot",
                side_effect=fake_export,
            ), patch(
                "TwitchChannelPointsMiner.TwitchChannelPointsMiner.interruptible_sleep",
                return_value=None,
            ):
                miner._streamers_export_loop()

            self.assertEqual(len(calls), 1)


if __name__ == "__main__":
    unittest.main()
